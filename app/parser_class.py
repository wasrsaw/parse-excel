import openpyxl

"""
Здесь основной функционал, мы парсим excel файл.
"""
class XParser:
    group_row = 2
    weekday_col = 1
    para_col = 2
    subgroups = ["1пг", "2пг"]
    parity = ["IIн", "Iн"]

    def __init__(self, path: str) -> None:
        self.PATH_ = path

    def get_weeks(self, input: str) -> list[str]:
        try:
            text = input.split()
            if text[1] == "н.": #Поскольку в расписании после указания недель идет 'н.'
                #проверяем его наличие
                return text[0].split(",")
        except Exception:
            #add logs
            pass
        finally:
            return None
        
    def get_week_parity(self, input: str) -> str:
        try:
            for i in self.parity:
                if i in input:
                    return i
        except TypeError:
            #add logs
            pass
        finally:
            return None
        
    def get_subgroup(self, input: str) -> str:
        try:
            for subgroup in self.subgroups:
                if subgroup in input:
                    return subgroup
        except TypeError:
            #add logs
            pass
        finally:
            return None

    def parse(self) -> dict:
        """
        Парсим данные
        """
        ws = openpyxl.load_workbook(self.PATH_, read_only=True).active

        #------------------Определяем границы расписания------------------
        min_col = ws.min_column
        min_row = ws.min_row
        max_col = ws.max_column
        max_row = ws.max_row
        #-----------------------------------------------------------------

        #---------------Парсим расписание в цикле по группам--------------
        groups = {}
        for col in range(min_col, max_col+1):
            group = ws.cell(self.group_row, col).value
            if group:  # Парсим по группам
                weekday_paras = {}

                for row0 in range(self.group_row+1, max_row):
                    weekday = ws.cell(row0, self.weekday_col).value
                    if weekday:  # Парсим по дню недели
                        weekday_paras[weekday] = []
                        for row1 in range(row0, max_row): 
                            check_day = ws.cell(row1, self.weekday_col).value
                            para_cell = ws.cell(row1, self.para_col).value
                            lesson_cell = ws.cell(row1, col).value
                            prep_cell = ws.cell(row1, col+1).value
                            
                            if check_day and check_day != weekday:
                                break

                            if para_cell:
                                order = para_cell
                                weekday_paras[weekday].append([])

                            if lesson_cell or prep_cell:
                                lesson = lesson_cell
                                prep = prep_cell
                                subgroup = self.get_subgroup(lesson_cell)
                                weeks = self.get_weeks(lesson_cell)
                                parity = self.get_week_parity(lesson_cell)

                                data = {
                                        "Пара": lesson, 
                                        "Преподаватель": prep,
                                        "Подгруппа": subgroup,
                                        "Недели": weeks,
                                        "Четность": parity,
                                        }
                                weekday_paras[weekday][order-1].append(data)
                                
                groups[group] = weekday_paras              
        return groups

class ParsedData:
    def txt_format(self):
        """
        В формате текста
        """
        pass

    def dict_format(self):
        """
        В формате словаря
        """
        pass

    def list_formt(self):
        """
        В формате списка
        """
        pass